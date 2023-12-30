from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
import json
from unidecode import unidecode
import inspect
from openpyxl import Workbook, load_workbook

# service = Service(executable_path="C:/chromedriver-win64/chromedriver.exe")   
# options = Options()
# options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
# driver = webdriver.Chrome(service=service, options=options)

driver = webdriver.Chrome()

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

workbooks = Workbook()
sheet = workbooks.active
items = ['Company name', 'CNPJ', 'Corporate reason', 'Street', 'City', 'Phone Number']

for id in range(0, 6):
    sheet.cell(row = 1, column = id + 1).value = items[id]

with open('stores.json', 'r') as file:
    stores = json.load(file)

store_error = []
count_file = 0
total_count = len(stores)
start_row = 2
for index, store in enumerate(stores):
    vendor_name = store['company_name']
    print(f'{total_count} -> {index + 1}')
    print(stores[index])
    print(vendor_name)
    sheet.cell(row=start_row, column=1).value = vendor_name
    url = 'https://www.magazineluiza.com.br/lojista/' + bytes(vendor_name, 'utf-8').decode('unicode_escape').encode('latin1').decode('utf-8').replace(' ', '').replace('.', '').lower()
    try:
        driver.get(url)
        print('done')
        product = driver.find_element(By.XPATH, '//*[@id="__next"]/div/main/section[4]/div[2]/div/ul/li[1]/a')
        driver.execute_script('arguments[0].click();', product)
        info_button = driver.find_element(By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[1]/p[1]/label')
        driver.execute_script('arguments[0].click();', info_button)
        print('info_button')
        cnpj = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[2]/div[2]/div[2]/div[3]/div[1]/p[2]').text
        print(cnpj)
        corporate_reason = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[2]/div[2]/div[2]/div[3]/div[2]/p[2]').text
        print(corporate_reason)
        street = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[2]/div[2]/div[2]/div[3]/div[3]/p[2]/p[1]').text
        print(street)
        city = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[2]/div[2]/div[2]/div[3]/div[3]/p[2]/p[2]').text
        print(city)
        phone_number = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[5]/div[3]/div[2]/div[2]/div[2]/div[3]/div[3]/p[2]/p[3]').text
        print(phone_number)
        sheet.cell(row=start_row, column=2).value = cnpj
        sheet.cell(row=start_row, column=3).value = corporate_reason
        sheet.cell(row=start_row, column=4).value = street
        sheet.cell(row=start_row, column=5).value = city
        sheet.cell(row=start_row, column=6).value = phone_number
    except:
        pass
    # driver.delete_all_cookies()
    workbooks.save('output.xlsx')
    start_row += 1

